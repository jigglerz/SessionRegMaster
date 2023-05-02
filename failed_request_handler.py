import openpyxl
from PyQt6.QtWidgets import QFileDialog


class FailedRequests:
    def __init__(self):
        self.failed_requests = {}

    def add_failed_request(self, session_id, ticket_id):
        if session_id not in self.failed_requests:
            self.failed_requests[session_id] = []

        self.failed_requests[session_id].append(ticket_id)

    def save_to_excel(self, parent_widget):
        if not self.failed_requests:  # Check if the dictionary is empty
            return

        file_name, _ = QFileDialog.getSaveFileName(parent_widget, "Save Failed Requests", "",
                                                   "Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            for idx, (key, values) in enumerate(self.failed_requests.items()):
                sheet.cell(row=1, column=idx + 1).value = key
                for row_idx, value in enumerate(values, start=2):
                    sheet.cell(row=row_idx, column=idx + 1).value = value

            workbook.save(file_name)
